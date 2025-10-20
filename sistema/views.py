from django.views import View
from django.shortcuts import render, redirect
from .models import Chamados
from django.contrib.auth.models import User
from datetime import datetime, date, timedelta
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import pandas as pd
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import os
from django.utils import timezone
from django.db import models



def dashboards(request):
    """
    View de dashboards - todos os filtros são aplicados no frontend.
    """
    # Buscar todos os chamados ordenados por data (mais recentes primeiro)
    chamados_qs = Chamados.objects.all().order_by('-data', '-id')

    # Converter queryset para lista serializável pelo json_script
    chamados_serializaveis = []
    for c in chamados_qs:
        chamados_serializaveis.append({
            'ID_chamado': c.ID_chamado or 0,
            'nome_analista': c.nome_analista.username if getattr(c, 'nome_analista', None) else 'N/A',
            'nome_tecnico': c.nome_tecnico if getattr(c, 'nome_tecnico', None) else 'N/A',
            'nome_cliente': c.nome_cliente if getattr(c, 'nome_cliente', None) else 'N/A',
            'tipo_atividade': c.tipo_atividade if getattr(c, 'tipo_atividade', None) else 'N/A',
            'inicio': c.inicio.strftime('%H:%M:%S') if getattr(c, 'inicio', None) else '00:00:00',
            'conclusao': c.conclusao.strftime('%H:%M:%S') if getattr(c, 'conclusao', None) else '00:00:00',
            'total_horas': str(c.total_horas) if getattr(c, 'total_horas', None) else '00:00:00',
            'produtiva': bool(c.produtiva) if getattr(c, 'produtiva', None) is not None else True,
            'data': c.data.strftime('%Y-%m-%d') if getattr(c, 'data', None) else '',
            'data_planejada': c.data_planejada.strftime('%Y-%m-%d') if getattr(c, 'data_planejada', None) else '',
            'status': c.status if getattr(c, 'status', None) else 'planejado',
            'senha': c.senha if getattr(c, 'senha', None) else 'N/A',
            'observacao': c.observacao if getattr(c, 'observacao', None) else 'N/A',
            'origem_planilha': bool(c.origem_planilha) if getattr(c, 'origem_planilha', None) is not None else False,
            'data_upload': c.data_upload.strftime('%Y-%m-%d %H:%M') if getattr(c, 'data_upload', None) else ''
        })

    return render(request, "dashboards.html", {
        "chamados": chamados_serializaveis
    })

def nomes_analistas(request):
    analistas_str = []
    analistas_object = User.objects.filter(is_superuser=False)
   
   
    for analista in analistas_object:
        nome_analista = User.objects.get(username=analista)

        analista = str(analista).replace("_"," ")
        analistas_str.append((analista,nome_analista.id))
    


    return render(request, "analistas.html", {"analistas": analistas_str })

def ver_analista(request, user_id):
    hora = datetime.strptime("01:00","%H:%M").time()

    analista = User.objects.get(id=user_id)
    hoje = date.today()
    chamados_feitos = Chamados.objects.filter(nome_analista=analista).filter(
        models.Q(data_planejada__lte=hoje) | models.Q(data_planejada__isnull=True)
    )
    analista = str(analista).replace("_", " ")

    return render(request, "chamado_especifico.html" , {"chamados": chamados_feitos, "analista": analista, "hora": hora} )


def todos_chamados(request):
    """
    View para exibir todos os chamados com suporte a filtros.
    Usa date.today() para obter a data atual do sistema.
    """
    # Obter data atual usando date.today()
    hoje = date.today()
    
    # Buscar todos os chamados ordenados por data (mais recentes primeiro)
    # Filtrar chamados que já podem ser exibidos (data_planejada <= hoje ou null)
    chamados = Chamados.objects.filter(
        models.Q(data_planejada__lte=hoje) | models.Q(data_planejada__isnull=True)
    ).order_by('-data', '-id')
    quantidade = chamados.count()
    
    # Filtros opcionais via GET parameters (para uso futuro)
    filtro_periodo = request.GET.get('period', None)
    
    if filtro_periodo == 'today':
        # Filtrar apenas chamados de hoje
        chamados = Chamados.objects.filter(data=hoje).order_by('-id')
        quantidade = chamados.count()
        print(f"🗓️ Filtro HOJE aplicado (todos_chamados): {hoje} - {quantidade} chamados")
    
    elif filtro_periodo == '7':
        # Últimos 7 dias
        data_limite = hoje - timedelta(days=7)
        chamados = Chamados.objects.filter(data__gte=data_limite).order_by('-data', '-id')
        quantidade = chamados.count()
    
    elif filtro_periodo == '30':
        # Últimos 30 dias
        data_limite = hoje - timedelta(days=30)
        chamados = Chamados.objects.filter(data__gte=data_limite).order_by('-data', '-id')
        quantidade = chamados.count()

    return render(request, "todos_chamados.html", {
        "chamados": chamados, 
        "quantidade": quantidade,
        "data_hoje": hoje  # Enviar data atual para o template
    })


def tabela_chamados(request):
    """
    View para exibir chamados em formato de tabela dinâmica (estilo Excel).
    Retorna todos os chamados ordenados por data decrescente para permitir
    manipulação completa no frontend (ordenação, filtros, busca, etc).
    """
    # Aplicar filtros da URL - suporte para 'period' (do dashboard) e 'periodo' (legado)
    periodo = request.GET.get('period', request.GET.get('periodo', ''))
    analista = request.GET.get('analista', '')
    tipo_atividade = request.GET.get('tipo_atividade', '')
    produtividade = request.GET.get('produtividade', '')
    status = request.GET.get('status', '')
    origem_planilha = request.GET.get('origem_planilha', '')
    
    # Construir query base - chamados "em_andamento" sempre no topo
    from django.db.models import Case, When, Value, IntegerField
    hoje = timezone.now().date()
    chamados = Chamados.objects.filter(
        models.Q(data_planejada__lte=hoje) | models.Q(data_planejada__isnull=True)
    ).order_by(
        Case(
            When(status='em_andamento', then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        ),
        '-data', 
        '-inicio'
    )
    
    # Aplicar filtros
    if periodo == 'hoje':
        hoje = timezone.now().date()
        chamados = chamados.filter(data=hoje)
    elif periodo == 'semana':
        semana_passada = timezone.now().date() - timedelta(days=7)
        chamados = chamados.filter(data__gte=semana_passada)
    elif periodo == 'mes':
        mes_passado = timezone.now().date() - timedelta(days=30)
        chamados = chamados.filter(data__gte=mes_passado)
    
    if analista:
        chamados = chamados.filter(nome_analista__username=analista)
    
    if tipo_atividade:
        chamados = chamados.filter(tipo_atividade=tipo_atividade)
    
    if produtividade == 'true':
        chamados = chamados.filter(produtiva=True)
    elif produtividade == 'false':
        chamados = chamados.filter(produtiva=False)
    
    if status:
        if status == 'produtiva':
            chamados = chamados.filter(produtiva=True)
        elif status == 'improdutiva':
            chamados = chamados.filter(produtiva=False)
        elif status == 'planejadas':
            chamados = chamados.filter(status='planejado')
        elif status == 'andamento':
            chamados = chamados.filter(status='em_andamento')
        elif status == 'finalizado':
            chamados = chamados.filter(status='finalizado')
    
    if origem_planilha == 'false':
        chamados = chamados.filter(origem_planilha=False)
    elif origem_planilha == 'true':
        chamados = chamados.filter(origem_planilha=True)
    
    analistas = User.objects.all()
    quantidade = chamados.count()
    
    # Retornar template com todos os dados
    return render(request, "tabela_chamados.html", {
        "chamados": chamados,
        "quantidade": quantidade,
        "analistas": analistas,
        "filtros": {
            "periodo": periodo,
            "analista": analista,
            "tipo_atividade": tipo_atividade,
            "produtividade": produtividade,
            "status": status,
            "origem_planilha": origem_planilha
        }
    })


@csrf_exempt
@require_http_methods(["POST"])
def salvar_dados_iniciais(request):
    """
    View para salvar dados iniciais do chamado via AJAX.
    Recebe dados JSON e retorna resposta JSON.
    """
    try:
        # Parsear dados JSON
        data = json.loads(request.body)
        
        # Validar dados obrigatórios
        campos_obrigatorios = ['nome_analista', 'ID_chamado', 'tipo_atividade', 'tecnico', 'data', 'inicio']
        for campo in campos_obrigatorios:
            if not data.get(campo):
                return JsonResponse({
                    'success': False,
                    'message': f'Campo obrigatório não preenchido: {campo}'
                }, status=400)
        
        # Validar se o analista existe
        try:
            analista = User.objects.get(username=data['nome_analista'])
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Analista não encontrado'
            }, status=400)
        
        # Verificar se o ID do chamado já existe
        chamado_existente = Chamados.objects.filter(ID_chamado=data['ID_chamado']).first()
        if chamado_existente:
            # Se o chamado existe e está planejado, atualizar para em_andamento
            if chamado_existente.status == 'planejado':
                # Atualizar o chamado planejado com os novos dados
                chamado_existente.nome_analista = analista
                chamado_existente.tipo_atividade = data['tipo_atividade']
                chamado_existente.nome_tecnico = data['tecnico']
                chamado_existente.data = datetime.strptime(data['data'], '%Y-%m-%d').date()
                chamado_existente.inicio = datetime.strptime(data['inicio'], '%H:%M').time()
                chamado_existente.status = 'em_andamento'
                chamado_existente.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Chamado assumido com sucesso! Status alterado para "Em Andamento".',
                    'chamado_existente': True,
                    'status_atual': 'em_andamento'
                })
            elif chamado_existente.status == 'em_andamento':
                return JsonResponse({
                    'success': True,
                    'message': 'Chamado em andamento encontrado. Pode finalizar o chamado.',
                    'chamado_existente': True,
                    'status_atual': 'em_andamento'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'Chamado já existe com status: {chamado_existente.get_status_display()}'
                }, status=400)
        
        # Validar formato da data
        try:
            datetime.strptime(data['data'], '%Y-%m-%d')
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Formato de data inválido'
            }, status=400)
        
        # Validar formato do horário
        try:
            datetime.strptime(data['inicio'], '%H:%M')
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Formato de horário inválido'
            }, status=400)
        
        # Salvar chamado com status "em_andamento"
        try:
            # Se o chamado já existe, atualizar
            if chamado_existente:
                chamado_existente.nome_analista = analista
                chamado_existente.tipo_atividade = data['tipo_atividade']
                chamado_existente.nome_tecnico = data['tecnico']
                chamado_existente.data = data['data']
                chamado_existente.inicio = data['inicio']
                chamado_existente.status = 'em_andamento'
                chamado_existente.origem_planilha = False
                chamado_existente.save()
                
                print(f"✅ Chamado {data['ID_chamado']} atualizado para 'em_andamento'")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Chamado atualizado e iniciado com sucesso!'
                })
            else:
                # Criar novo chamado
                novo_chamado = Chamados.objects.create(
                    nome_analista=analista,
                    ID_chamado=data['ID_chamado'],
                    tipo_atividade=data['tipo_atividade'],
                    nome_tecnico=data['tecnico'],
                    data=data['data'],
                    inicio=data['inicio'],
                    status='em_andamento',
                    origem_planilha=False
                )
                
                print(f"✅ Novo chamado {data['ID_chamado']} criado com status 'em_andamento'")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Chamado iniciado com sucesso! Agora ele aparece na tabela.'
                })
                
        except Exception as e:
            print(f"❌ Erro ao salvar chamado: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Erro ao salvar chamado: {str(e)}'
            }, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Dados JSON inválidos'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro interno: {str(e)}'
        }, status=500)




def views(request):
    

    if request.method == "POST":

        formato = "%Y-%m-%d"
        data_especifica = request.POST.get("data") 
        data_especifica = datetime.strptime(data_especifica, formato).date()
        data_hoje = date.today()



        hora = datetime.strptime("01:00","%H:%M").time()

        

        hoje = date.today()
        chamados = Chamados.objects.filter(data=data_especifica).filter(
            models.Q(data_planejada__lte=hoje) | models.Q(data_planejada__isnull=True)
        )
        quantidade = chamados.count() 

        return render(request, "visualização.html", {"chamados":chamados, "quantidade": quantidade, "data_especificada":data_especifica,"data_hoje": data_hoje, "hora":hora}  )
    else:


        hora = datetime.strptime("01:00", "%H:%M").time()
        data_especifica = date.today()
        data_hoje = date.today()



        hoje = date.today()
        chamados = Chamados.objects.filter(
            models.Q(data_planejada__lte=hoje) | models.Q(data_planejada__isnull=True)
        )
        quantidade = chamados.count()
        return render(request, "visualização.html",  {"chamados":chamados, "quantidade": quantidade,"data_especificada":data_especifica,"data_hoje":data_hoje, "hora":hora})



class RegistrarChamado(View):

    def get(self, request):
        return render(request, "index.html")
    

    def post(self, request):
        try:
            # Coletar dados do formulário
            self.nome_analista = request.POST.get("nome_analista")
            self.ID_chamado = request.POST.get("ID_chamado")
            self.tipo_atividade = request.POST.get("tipo_atividade")
            self.nome_tecnico = request.POST.get("tecnico")
            self.data = request.POST.get("data")
            self.inicio = request.POST.get("inicio")
            self.conclusao = request.POST.get("conclusao")
            self.situacao = request.POST.get("produtiva")
            self.senha = request.POST.get("senha")
            self.observacao = request.POST.get("observacao")

            # Validar dados obrigatórios
            if not all([self.nome_analista, self.ID_chamado, self.tipo_atividade, 
            self.nome_tecnico, self.data, self.inicio, self.conclusao, 
            self.senha]):
                return render(request, "index.html", {
                    'error': 'Todos os campos obrigatórios devem ser preenchidos.'
                })

            # Processar dados
            RegistrarChamado._validar_situacao(self)
            RegistrarChamado._cauculo_de_tempo_de_atendimento(self)
            RegistrarChamado._salvador_chamado(self)
            
            # Sucesso - redirecionar com mensagem
            return render(request, "index.html", {
                'success': 'Chamado registrado com sucesso!'
            })
            
        except Exception as e:
            print(f"Erro no registro de chamado: {str(e)}")
            return render(request, "index.html", {
                'error': f'Erro ao registrar chamado: {str(e)}'
            })

    


    def _validar_situacao(self):
        if self.situacao == "on":

            self.situacao = True
        
        else:
            self.situacao = False
   
        
    def _cauculo_de_tempo_de_atendimento(self):
        formato = "%H:%M"

        inicio = datetime.strptime(self.inicio, formato)
        conclusao = datetime.strptime(self.conclusao, formato)
        
        # Calcular diferença em horas (float)
        diferenca = conclusao - inicio
        self.total_horas = diferenca.total_seconds() / 3600.0
        
        # Se o tempo for 0 ou negativo, considerar 1 minuto mínimo
        if self.total_horas <= 0:
            self.total_horas = 1/60.0  # 1 minuto = 1/60 de hora
    

    def _salvador_chamado(self):
        try:
            print(f"Salvando chamado: {self.tipo_atividade}")

            # Verificar se o chamado já existe (vindo de planilha)
            chamado_existente = Chamados.objects.filter(ID_chamado=self.ID_chamado).first()
            
            if chamado_existente:
                # Chamado já existe - atualizar status e dados
                print(f"Chamado {self.ID_chamado} já existe. Atualizando status...")
                
                # Determinar novo status baseado no status atual
                if chamado_existente.status == 'planejado':
                    novo_status = 'em_andamento'
                    print(f"Status alterado de 'planejado' para 'em_andamento'")
                elif chamado_existente.status == 'em_andamento':
                    novo_status = 'finalizado'
                    print(f"Status alterado de 'em_andamento' para 'finalizado'")
                else:
                    novo_status = chamado_existente.status  # Manter status atual
                
                # Atualizar chamado existente
                chamado_existente.nome_analista = User.objects.get(username=self.nome_analista)
                chamado_existente.tipo_atividade = self.tipo_atividade
                chamado_existente.nome_tecnico = self.nome_tecnico
                chamado_existente.data = self.data
                chamado_existente.inicio = self.inicio
                chamado_existente.conclusao = self.conclusao
                chamado_existente.total_horas = self.total_horas
                chamado_existente.produtiva = self.situacao
                chamado_existente.senha = self.senha
                chamado_existente.observacao = self.observacao
                chamado_existente.status = novo_status
                chamado_existente.save()
                
                print(f"Chamado {self.ID_chamado} atualizado com sucesso! Status: {novo_status}")
            else:
                # Chamado não existe - criar novo
                print(f"Criando novo chamado {self.ID_chamado}")

                Chamados.objects.create(
                    nome_analista = User.objects.get(username=self.nome_analista),
                    ID_chamado = self.ID_chamado,
                    tipo_atividade = self.tipo_atividade,
                    nome_tecnico = self.nome_tecnico,
                    data = self.data,
                    inicio = self.inicio,
                    conclusao = self.conclusao,
                    total_horas = self.total_horas,
                    produtiva = self.situacao,
                    senha = self.senha,
                    observacao = self.observacao,
                    status = 'em_andamento',  # Novo chamado começa como "em andamento"
                    origem_planilha = False
                )
                print("Novo chamado criado com sucesso!")
                
        except Exception as e:
            print(f"Erro ao salvar chamado: {str(e)}")
            raise e


def exportar_excel_formatado(request):
    """
    View para exportar chamados em formato Excel com formatação personalizada.
    Aplica as cores e estilos do site (dourado e escuro).
    """
    try:
        # Buscar todos os chamados ordenados por data (mais recentes primeiro)
        chamados = Chamados.objects.all().order_by('-data', '-id')
        
        # Criar workbook e worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Chamados"
        
        # ========== CONFIGURAÇÕES DE ESTILO (CORES DO SITE) ==========
        
        # Cores do site
        COR_DOURADO = "F8C24A"  # Dourado principal do site
        COR_ESCURO = "2A2A2A"   # Escuro principal do site
        COR_ESCURO_CLARO = "1F1F1F"  # Escuro mais claro
        COR_CINZA_CLARO = "CCCCCC"   # Cinza claro para texto
        COR_VERDE = "4ADE80"    # Verde para produtiva
        COR_VERMELHO = "EF4444" # Vermelho para improdutiva
        
        # Estilos de fonte
        fonte_cabecalho = Font(
            name='Inter',
            size=12,
            bold=True,
            color=COR_DOURADO
        )
        
        fonte_dados = Font(
            name='Inter',
            size=11,
            color=COR_CINZA_CLARO
        )
        
        fonte_produtiva = Font(
            name='Inter',
            size=11,
            bold=True,
            color=COR_VERDE
        )
        
        fonte_improdutiva = Font(
            name='Inter',
            size=11,
            bold=True,
            color=COR_VERMELHO
        )
        
        # Preenchimentos (cores de fundo)
        preenchimento_cabecalho = PatternFill(
            start_color=COR_ESCURO,
            end_color=COR_ESCURO,
            fill_type="solid"
        )
        
        preenchimento_linha_par = PatternFill(
            start_color=COR_ESCURO_CLARO,
            end_color=COR_ESCURO_CLARO,
            fill_type="solid"
        )
        
        preenchimento_linha_impar = PatternFill(
            start_color=COR_ESCURO,
            end_color=COR_ESCURO,
            fill_type="solid"
        )
        
        # Alinhamento
        alinhamento_centro = Alignment(horizontal='center', vertical='center')
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
        
        # Bordas
        borda_fina = Side(style='thin', color=COR_DOURADO)
        borda = Border(
            left=borda_fina,
            right=borda_fina,
            top=borda_fina,
            bottom=borda_fina
        )
        
        # ========== CABEÇALHOS ==========
        cabecalhos = [
            'Analista', 'ID Chamado', 'Tipo Atividade', 'Técnico',
            'Data', 'Horário Previsto', 'Início', 'Conclusão', 'Tempo Total',
            'Status', 'Senha', 'Observação'
        ]
        
        # Aplicar cabeçalhos com formatação
        for col, cabecalho in enumerate(cabecalhos, 1):
            cell = worksheet.cell(row=1, column=col, value=cabecalho)
            cell.font = fonte_cabecalho
            cell.fill = preenchimento_cabecalho
            cell.alignment = alinhamento_centro
            cell.border = borda
        
        # ========== DADOS ==========
        for row, chamado in enumerate(chamados, 2):
            # Determinar cor de fundo (alternando linhas)
            preenchimento = preenchimento_linha_par if row % 2 == 0 else preenchimento_linha_impar
            
            # Dados da linha
            dados_linha = [
                chamado.nome_analista.username if chamado.nome_analista else '',
                chamado.ID_chamado,
                chamado.tipo_atividade,
                chamado.nome_tecnico,
                chamado.data.strftime('%d/%m/%Y') if chamado.data else '',
                chamado.previsto.strftime('%H:%M') if chamado.previsto else '',
                chamado.inicio.strftime('%H:%M') if chamado.inicio else '',
                chamado.conclusao.strftime('%H:%M') if chamado.conclusao else '',
                chamado.total_horas.strftime('%H:%M') if chamado.total_horas else '',
                'Produtiva' if chamado.produtiva else 'Improdutiva',
                chamado.senha,
                chamado.observacao or 'Sem observações'
            ]
            
            # Aplicar dados com formatação
            for col, dado in enumerate(dados_linha, 1):
                cell = worksheet.cell(row=row, column=col, value=dado)
                cell.fill = preenchimento
                cell.border = borda
                cell.alignment = alinhamento_esquerda
                
                # Fonte especial para status
                if col == 9:  # Coluna Status
                    if chamado.produtiva:
                        cell.font = fonte_produtiva
                    else:
                        cell.font = fonte_improdutiva
                else:
                    cell.font = fonte_dados
        
        # ========== AJUSTAR LARGURA DAS COLUNAS ==========
        larguras_colunas = [15, 12, 20, 12, 12, 15, 10, 12, 12, 12, 12, 30]
        
        for col, largura in enumerate(larguras_colunas, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = largura
        
        # ========== CONGELAR PRIMEIRA LINHA ==========
        worksheet.freeze_panes = 'A2'
        
        # ========== PREPARAR RESPOSTA ==========
        # Criar buffer em memória
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        # Nome do arquivo com data/hora atual
        agora = datetime.now()
        nome_arquivo = f"chamados_formatados_{agora.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Criar resposta HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        print(f"✅ Excel formatado exportado: {nome_arquivo}")
        return response
        
    except Exception as e:
        print(f"❌ Erro ao exportar Excel: {str(e)}")
        return JsonResponse({'error': f'Erro ao exportar Excel: {str(e)}'}, status=500)


def upload_planilha(request):
    """
    View para upload e processamento de planilhas de chamados planejados.
    Acesso permitido apenas das 16:10 às 17:10, uma vez por dia.
    """
    # Verificar horário de acesso (16:10 às 17:10) - usando horário local
    agora = datetime.now()  # Usar horário local ao invés de timezone.now()
    hora_atual = agora.time()
    hora_inicio = datetime.strptime('16:10', '%H:%M').time()
    hora_fim = datetime.strptime('17:10', '%H:%M').time()
    
    # Verificar se já foi feito upload hoje
    hoje = agora.date()
    uploads_hoje = Chamados.objects.filter(data_upload__date=hoje, origem_planilha=True).count()
    
    # Permitir bypass apenas para hoje com parâmetro especial
    permitir_novo_upload = request.GET.get('permitir_novo', '') == 'sim'
    
    # Permitir segundo upload hoje se estiver no horário correto (13:20-16:00)
    dentro_do_horario = hora_inicio <= hora_atual <= hora_fim
    
    # Verificar se já foi feito upload hoje (só bloqueia se estiver FORA do horário)
    if uploads_hoje > 0 and not permitir_novo_upload and not dentro_do_horario:
        if request.method == 'GET':
            return render(request, 'upload_planilha.html', {
                'horario_restrito': True,
                'hora_inicio': '16:10',
                'hora_fim': '17:10',
                'ja_upload_hoje': True
            })
        else:
            return JsonResponse({
                'error': 'Upload já realizado hoje. Faça novos uploads apenas das 16:10 às 17:10.'
            }, status=403)
    
    # Verificar horário (mas permitir bypass se tiver o parâmetro)
    if not (hora_inicio <= hora_atual <= hora_fim) and not permitir_novo_upload:
        if request.method == 'GET':
            return render(request, 'upload_planilha.html', {
                'horario_restrito': True,
                'hora_inicio': '16:10',
                'hora_fim': '17:10',
                'ja_upload_hoje': False
            })
        else:
            return JsonResponse({
                'error': f'Upload de planilha permitido apenas das {hora_inicio.strftime("%H:%M")} às {hora_fim.strftime("%H:%M")}'
            }, status=403)
    
    if request.method == 'GET':
        return render(request, 'upload_planilha.html', {
            'horario_restrito': False,
            'hora_inicio': '16:10',
            'hora_fim': '17:10'
        })
    
    if request.method == 'POST':
        try:
            # Verificar se arquivo foi enviado
            if 'planilha' not in request.FILES:
                return JsonResponse({'error': 'Nenhum arquivo foi enviado'}, status=400)
            
            arquivo = request.FILES['planilha']
            
            # Verificar extensão do arquivo
            if not arquivo.name.endswith(('.xlsx', '.xls', '.csv')):
                return JsonResponse({'error': 'Formato de arquivo não suportado. Use .xlsx, .xls ou .csv'}, status=400)
            
            # Ler planilha
            if arquivo.name.endswith('.csv'):
                # Tentar diferentes encodings e delimitadores para CSV
                df = None
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                separators = [',', ';', '\t']
                
                for encoding in encodings:
                    for sep in separators:
                        try:
                            arquivo.seek(0)  # Resetar posição do arquivo
                            df = pd.read_csv(arquivo, encoding=encoding, sep=sep)
                            # Verificar se conseguiu ler pelo menos uma coluna
                            if len(df.columns) > 0:
                                break
                        except:
                            continue
                    if df is not None and len(df.columns) > 0:
                        break
                
                if df is None or len(df.columns) == 0:
                    return JsonResponse({
                        'error': 'Não foi possível ler o arquivo CSV. Verifique se o arquivo não está corrompido e se tem o formato correto.'
                    }, status=400)
            else:
                df = pd.read_excel(arquivo)
            
            # Mapeamento de colunas alternativas
            mapeamento_colunas = {
                'TICKET': 'ID_chamado',
                'Ponto': 'nome_cliente', 
                'TÉCNICO': 'nome_tecnico',
                'DATA': 'data_planejada',
                'SERVIÇO': 'tipo_atividade',
                'Horário Previsto': 'previsto'
            }
            
            # Renomear colunas se necessário
            print(f"📊 Colunas detectadas: {list(df.columns)}")
            for col_original, col_nova in mapeamento_colunas.items():
                if col_original in df.columns and col_nova not in df.columns:
                    df = df.rename(columns={col_original: col_nova})
                    print(f"🔄 Coluna renomeada: {col_original} -> {col_nova}")
            
            print(f"📊 Colunas após mapeamento: {list(df.columns)}")
            
            # Verificar colunas obrigatórias
            colunas_obrigatorias = ['ID_chamado', 'nome_cliente', 'nome_tecnico', 'data_planejada']
            colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]
            
            if colunas_faltando:
                # Mostrar colunas disponíveis para ajudar o usuário
                colunas_disponiveis = list(df.columns)
                return JsonResponse({
                    'error': f'Colunas obrigatórias não encontradas: {", ".join(colunas_faltando)}. Colunas disponíveis: {", ".join(colunas_disponiveis)}'
                }, status=400)
            
            # Processar dados
            chamados_criados = 0
            chamados_duplicados = 0
            erros = []
            
            for index, row in df.iterrows():
                try:
                    # Verificar se chamado já existe
                    if Chamados.objects.filter(ID_chamado=row['ID_chamado']).exists():
                        chamados_duplicados += 1
                        continue
                    
                    # Converter data planejada
                    data_planejada = None
                    if pd.notna(row['data_planejada']):
                        if isinstance(row['data_planejada'], str):
                            # Tentar diferentes formatos de data
                            try:
                                data_planejada = datetime.strptime(row['data_planejada'], '%d/%m/%Y').date()
                            except:
                                try:
                                    data_planejada = datetime.strptime(row['data_planejada'], '%Y-%m-%d').date()
                                except:
                                    try:
                                        data_planejada = datetime.strptime(row['data_planejada'], '%Y/%m/%d').date()
                                    except:
                                        print(f"❌ Formato de data não reconhecido: {row['data_planejada']}")
                                        raise ValueError(f"Formato de data não reconhecido: {row['data_planejada']}")
                        else:
                            data_planejada = row['data_planejada'].date()
                    
                    # Converter horário previsto
                    horario_previsto = None
                    print(f"🕐 Verificando horário previsto para linha {index + 2}: {row.get('previsto', 'NÃO ENCONTRADO')}")
                    if 'previsto' in df.columns and pd.notna(row.get('previsto')):
                        print(f"🕐 Campo previsto encontrado: {row.get('previsto')}")
                        if isinstance(row['previsto'], str):
                            # Tentar diferentes formatos de horário
                            try:
                                horario_previsto = datetime.strptime(row['previsto'], '%H:%M').time()
                                print(f"🕐 Horário convertido (HH:MM): {horario_previsto}")
                            except:
                                try:
                                    horario_previsto = datetime.strptime(row['previsto'], '%H:%M:%S').time()
                                    print(f"🕐 Horário convertido (HH:MM:SS): {horario_previsto}")
                                except:
                                    print(f"❌ Não foi possível converter horário: {row['previsto']}")
                                    horario_previsto = None
                        else:
                            horario_previsto = row['previsto']
                            print(f"🕐 Horário já no formato correto: {horario_previsto}")
                    else:
                        print(f"❌ Campo 'previsto' não encontrado ou vazio")
                    
                    # Criar chamado
                    chamado = Chamados.objects.create(
                        ID_chamado=str(row['ID_chamado']),  # Usar string para suportar IDs como "125978/1"
                        nome_cliente=str(row['nome_cliente']),
                        nome_tecnico=str(row['nome_tecnico']),
                        data_planejada=data_planejada,
                        previsto=horario_previsto,
                        status='planejado',
                        origem_planilha=True,
                        data_upload=timezone.now(),
                        # Campos opcionais da planilha
                        tipo_atividade=str(row.get('tipo_atividade', '')),
                        observacao=str(row.get('observacao', ''))
                    )
                    
                    chamados_criados += 1
                    
                except Exception as e:
                    print(f"❌ Erro na linha {index + 2}: {str(e)}")
                    print(f"Dados da linha: {dict(row)}")
                    erros.append(f"Linha {index + 2}: {str(e)}")
            
            # Resposta de sucesso
            mensagem = f"Upload concluído! {chamados_criados} chamados criados."
            if chamados_duplicados > 0:
                mensagem += f" {chamados_duplicados} chamados duplicados ignorados."
            if erros:
                mensagem += f" {len(erros)} erros encontrados."
            
            return JsonResponse({
                'success': True,
                'message': mensagem,
                'chamados_criados': chamados_criados,
                'chamados_duplicados': chamados_duplicados,
                'erros': erros[:10]  # Limitar a 10 erros
            })
            
        except Exception as e:
            print(f"❌ Erro no upload: {str(e)}")
            return JsonResponse({'error': f'Erro ao processar planilha: {str(e)}'}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def finalizar_chamado(request):
    """
    View para finalizar um chamado.
    Atualiza os dados de conclusão, produtividade, senha e observação.
    """
    try:
        data = json.loads(request.body)
        
        ID_chamado = data.get('ID_chamado')
        conclusao = data.get('conclusao')
        produtiva = data.get('produtiva')
        senha = data.get('senha')
        observacao = data.get('observacao', '')
        
        # Validações
        if not ID_chamado:
            return JsonResponse({'success': False, 'message': 'ID do chamado não informado'}, status=400)
        
        if not conclusao:
            return JsonResponse({'success': False, 'message': 'Horário de conclusão não informado'}, status=400)
        
        # Converter produtividade para boolean
        if produtiva in ['true', 'True', '1', 'sim', 'Sim']:
            produtiva = True
        elif produtiva in ['false', 'False', '0', 'não', 'nao', 'Não', 'Nao']:
            produtiva = False
        else:
            return JsonResponse({'success': False, 'message': 'Produtividade inválida'}, status=400)
        
        # Buscar chamado
        try:
            chamado = Chamados.objects.get(ID_chamado=ID_chamado)
        except Chamados.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Chamado não encontrado'}, status=404)
        
        # Atualizar chamado
        chamado.conclusao = conclusao
        chamado.produtiva = produtiva
        chamado.senha = senha
        chamado.observacao = observacao
        chamado.status = 'finalizado'
        
        # Calcular total de horas se possível
        if chamado.inicio and conclusao:
            try:
                # chamado.inicio já é um objeto time, não precisa converter
                inicio_time = chamado.inicio
                conclusao_time = datetime.strptime(conclusao, '%H:%M').time()
                
                inicio_datetime = datetime.combine(date.today(), inicio_time)
                conclusao_datetime = datetime.combine(date.today(), conclusao_time)
                
                # Se conclusão for menor que início, assumir que passou para o dia seguinte
                if conclusao_datetime < inicio_datetime:
                    conclusao_datetime += timedelta(days=1)
                
                diferenca = conclusao_datetime - inicio_datetime
                horas = diferenca.total_seconds() / 3600.0
                
                # Se o tempo for 0 ou negativo, considerar 1 minuto mínimo
                if horas <= 0:
                    horas = 1/60.0  # 1 minuto = 1/60 de hora
                
                chamado.total_horas = horas
            except:
                pass  # Se houver erro no cálculo, manter o valor existente
        
        chamado.save()
        
        print(f"✅ Chamado {ID_chamado} finalizado com sucesso!")
        
        return JsonResponse({
            'success': True,
            'message': 'Chamado finalizado com sucesso!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Dados JSON inválidos'}, status=400)
    except Exception as e:
        print(f"❌ Erro ao finalizar chamado: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Erro ao finalizar chamado: {str(e)}'}, status=500)


